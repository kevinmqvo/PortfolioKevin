from openpyxl import Workbook
from openpyxl import load_workbook


wb = load_workbook(filename = "exemple-import-export-fois-10.xlsx")

ws = wb.active
for i in range(2,3):
    for z in range(2,25):
        a = ws.cell(row=z, column=i).value
        ws.cell(row=(z), column=(i+2)).value =(a*10)

wb.save("exemple-import-export-fois-10.xlsx")


