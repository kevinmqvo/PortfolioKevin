from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename = "exemple-import-export-fois-10.xlsx")
ws = wb.active

for i in range(2, 26):
    for j in range(2, 4):
        ws.cell(i,j+3).value = ws.cell(i,j).value * 10

wb.save("exemple-import-export-fois-10-out.xlsx")

