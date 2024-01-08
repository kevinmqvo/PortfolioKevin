from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename = "courbe")
ws = wb.active

for i in range(0,10):
    for j in range(0, 10):
        ws.cell(i,j)value =t[i][j]

wb.save("courbe-out")

