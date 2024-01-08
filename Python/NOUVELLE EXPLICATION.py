#On a écrit un introduction
print("DUOPOLE DE COURNOT: Deux entreprises se font concurrence sur la quantité du bien qu'elles produisent, le prix du marché étant le même pour les deux. Chacune d'elles a un coût de production unitaire propre:c1 pour E1 et c2 pour E2. Leur coût de production est donc: C1=c1*q1 pour E1 et c2=c2*q2 pour E2 où q1 et q2 sont les quantités produites respectivement par E1 et E2.")
print()
print("Le prix du marché est fixé par la fonction de demande inverse: P(q1,q2)=15-(q1+q2)")
print()
print()
#On rentre les variables exogènes de notre modèle, ici le coût unitaire à la production d'un bien chez chaque entreprise.
c1=int(input("entrez le coût unitaire de l'entreprise 1:"))
c2=int(input("entrez le coût unitaire de l'entreprise 2:"))
print()

#On crée un tableau T pour l'entreprise 1 avec 10 lignes et on insère des 0 dans les cases vides puis à l'aide  d'une boucle for on rajoute 10 colonnes pour chaque ligne.
t=[0]*10
for i in range(0,10):
    t[i]=[0]*10

z=[0]*10
#On crée un tableau Z pour l'entreprise 2 avec 10 lignes et on insère des 0 dans les cases vides puis à l'aide d'une boucle for on rajoute 10 colonnes pour chaque ligne.
for j in range(0,10):
    z[j]=[0]*10

#On calcule le profit que font les entreprises selon les quantités produites par l'entreprise 1 et par l'entreprise 2.
#On utilise une double boucle for car nous avons une combinaison de deux variables. La variable q1 et la variable q2.
#On calcule à l'aide de la fonction de profit (15-q)*q1-c1*q1
for q1 in range(0,10):
    for q2 in range(0,10):
        q=q1+q2
        pro=(15-q)*q1-c1*q1
        t[q1][q2]=pro

        q=q1+q2
        prof=(15-q)*q2-c2*q2
        z[q1][q2]=prof
#On insère les valeurs de profit obtenues dans le tableau T pour l'entreprise 1 et le tableau Z pour l'entreprise 2 et on affiche les tableaux. 
print(t)
print()
print(z)
print()

#On crée un tableau S de 10 cases pour y insérer les meilleures réponses de l'entreprise 1 et un tableau D pour y insérer les meilleures réponses de l'entreprise 2
s=[0]*10
d=[0]*10
k=0
#Dans la première boucle, on fixe q1 la quantité produite de l'entreprise 1. Dans la seconde boucle, on fait défiler les q2. On observe le profit Z obtenu. 
#La condition if permet de savoir si le profit Z sélectionné par la boucle est supérieur au profit précédent. Si il est supérieur, le profit est sauvegardé dans la variable K
#puis transféré dans la variable S.
#On sauvegarde aussi la valeur de q2 associée à ce profit dans la variable D.
#A la fin de chaque itération de la seconde boucle, on obtient le profit maximum S et la quantité optimale D pour un q1 donné.
#On a la quantité optimale que doit produire l'entreprise 2 selon les quantités de q1. Il s'agit de la meilleure réponse.

for q1 in range(0,10):
    for q2 in range(0,10):
        if z[q1][q2]>=k:
            k=z[q1][q2]
            s[q1]=k
            d[q1]=q2
    k=0

#Dans la première boucle, on fixe q2 la quantité produite de l'entreprise 2. Dans la seconde boucle, on fait défiler les q1. On observe le profit T obtenu. 
#La condition if permet de savoir si le profit T sélectionné par la boucle est supérieur au profit précédent. Si il est supérieur, le profit est sauvegardé dans la variable R
#puis transféré dans la variable F.
#On sauvegarde aussi la valeur de q1 associée à ce profit dans la variable F.
#A la fin de chaque itération de la seconde boucle, on obtient le profit maximum M et la quantité optimale F pour un  q2 donné.
#On a la quantité optimale que doit produire l'entreprise 1 selon les quantités de q2. Il s'agit de la meilleure réponse de l'entreprise 2.
m=[0]*10
f=[0]*10
r=0
for q2 in range (0,10):
    for q1 in range(0,10):
        if t[q1][q2]>=r:
            r=t[q1][q2]
            m[q2]=r
            f[q2]=q1
    r=0
#On affiche les meilleures réponses. D'abord celle de l'entreprise 1 puis l'entreprise 2.
print("prof1:",m)
print("q1*  :",f)
print()
print("prof2:",s)
print("q2*  :",d)
print()

#On fixe q2 dans la première boucle.
#Dans la seconde boucle, pour trouver l'equilibre, on fait défiler la meilleure réponse de l'entreprise 1 dans la variable g pour chaque q2 .
#La condition IF détermine si la meilleure réponse est un équilibre.
#Il s'agit d'un equilibre si: à une quantité q2 fixé par la première boucle, on regarde les meilleures réponses. L'entreprise 2 produit telle quantité à un certain q1.
#Et si cette quantité q1 est la même que celle sauvegardée en g,càd une meilleure réponse de l'entreprise 1, il s'agit d'une meilleure réponse.
#En effet cela veut dire que les deux q1 et q2 sélectionnés sont des meilleures réponses à la fois pour l'entreprise 1 et à la fois pour l'entreprise 2.


for q2 in range(0,10):
    for q1 in range(0,10):
        g=f[q2]
        if d[q1]==q2 and q1==g:
            sol2=q2
            sol1=g
#Nous avons remarqué que dans certains cas nous avions deux équilibres. Nous avons alors rajouté une seconde double boucle for
#où cette fois nous avons la condition où les résultats trouvés doivent être différents de ceux précedemment trouvés.
solt1=0
solt2=0
for q2 in range(0,10):
    for q1 in range(0,10):
        g=f[q2]
        if d[q1]==q2 and q1==g and q2!=sol2 and q1!=sol1:
            solt2=q2
            solt1=g
#Si aucun résultat n'a été trouvé, alors il n'y a qu'un seul équilibre donc on supprime les variables nouvellement crée.
if solt1==0 and solt2==0:
    solt1=sol1
    solt2=sol2
            
print()
print(sol1,sol2,solt1,solt2)
#Avec une condition IF, on a ajouté deux cas, l'un avec un seul équilibre et l'autre avec deux équilibres.
if sol1==solt1 and sol2==solt2:                
    print("A l'equilibre de cournot, les quantités sont:q1*=",sol1,"et q2*=",sol2,".L'entreprise 1 fait un profit de",t[sol1][sol2]," et l'entreprise 2 fait un profit de",z[sol1][sol2])
else:
    print("À l'equilibre,l'entreprise 1 produit",sol1,"et fait un profit de",t[sol1][sol2],". L'entreprise 2 produit",sol2,"et fait un profit de",z[sol1][sol2],". Au second equilibre, l'entreprise 1 produit",solt1," et fait un profit de",t[solt1][solt2],". L'entreprise 2 produit",solt2,"et fait un profit de",z[solt1][solt2],".")
#On a ensuite essayé d'exporter notre programme sur excel.

from openpyxl import Workbook
from openpyxl import load_workbook

#On sélectionne la feuille excel nommée "courbe" que l'on a préalablement sauvegardée dans notre dossier Python. 
wb = load_workbook(filename = "courbe.xlsx")
ws = wb.active
#A l'aide d'une boucle for, on fait défiler les lignes de la feuille excel à partir de la seconde ligne car nous avons
#mis les entêtes de tableau sur la première ligne.
#On insère les meilleures réponses de l'entreprise 1 dans la colonne 1 pour chaque ligne et celle de l'entreprise 2 dans la colonne D.
for i in range(0,10):
        ws.cell(i+2,2).value=f[i]
        ws.cell(i+2,5).value=d[i]
#La feuille résultante est sauvegardée sous le nom "courbe-out".
#On a une courbe affichant les meilleures réponses de l'entreprise 1 en bleu et celle l'entreprise 2 en vert.
#On remarque que le courbes se croisent au même points que l'équilibre trouvé par le programme.
wb.save("courbe-out.xlsx")




