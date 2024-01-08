print("DUOPOLE DE COURNOT: Deux entreprises se font concurrence sur la quantité du bien qu'elles produisent, le prix du marché étant le même pour les deux. Chacune d'elles a un coût de production unitaire propre:c1 pour E1 et c2 pour E2. Leur coût de production est donc: C1=c1*q1 pour E1 et c2=c2*q2 pour E2 où q1 et q2 sont les quantités produites respectivement par E1 et E2.")
print()
print("Le prix du marché est fixé par la fonction de demande inverse: P(q1,q2)=15-(q1+q2)")
print()
print()
c1=int(input("entrez le coût unitaire de l'entreprise 1:"))
c2=int(input("entrez le coût unitaire de l'entreprise 2:"))
print()

t=[0]*10
for i in range(0,10):
    t[i]=[0]*10

z=[0]*10
for j in range(0,10):
    z[j]=[0]*10

for q1 in range(0,10):
    for q2 in range(0,10):
        q=q1+q2
        pro=(15-q)*q1-c1*q1
        t[q1][q2]=pro
        
        q=q1+q2
        prof=(15-q)*q2-c2*q2
        z[q1][q2]=prof
print(t)
print()
print(z)
print()

s=[0]*10
d=[0]*10
k=0
for q1 in range(0,10):
    for q2 in range(0,10):
        if z[q1][q2]>=k:
            k=z[q1][q2]
            s[q1]=k
            d[q1]=q2
    k=0


m=[0]*10
f=[0]*10
r=0
for q2 in range (0,10):
    for q1 in range(0,10):
        if t[q1][q2]>=r:
            r=t[q1][q2]
            m[q2]=r
            f[q2]=q1
    r=0

print("prof1:",m)
print("q1*  :",f)
print()
print("prof2:",s)
print("q2*  :",d)
print()


for q2 in range(0,10):
    for q1 in range(0,10):
        g=f[q2]
        if d[q1]==q2 and q1==g:
            sol2=q2
            sol1=g
                
print("A l'equilibre de cournot, les quantités sont:q1*=",sol1,"et q2*=",sol2,".L'entreprise 1 fait un profit de",t[sol1][sol2]," et l'entreprise 2 fait un profit de",z[sol1][sol2])
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename = "courbe.xlsx")
ws = wb.active

for i in range(0,10):
        ws.cell(i+2,2).value=f[i]
        ws.cell(i+2,5).value=d[i]

wb.save("courbe-out.xlsx")




